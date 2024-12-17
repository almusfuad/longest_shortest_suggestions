import datetime
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import load_workbook


def get_day_of_week():
      """Get the name of the day of the week.

      Returns:
          String: Returns the name of the day of the week.
      """      
      return datetime.datetime.now().strftime("%A")





def read_keywords_from_excel(file_path, day_name):
      """Reads the keywords from the excel file and returns them along with the workbook and sheet objects.

      Args:
          file_path (String): Get the file path of the excel file.
          day_name (String): Get the name of the day to read the keywords from.

      Raises:
          ValueError: If the sheet with the given name is not found in the workbook.

      Returns:
          List: Returns the keywords, workbook, and sheet objects.
      """      
      workbook = load_workbook(file_path)
      
      if day_name not in workbook.sheetnames:
            raise ValueError(f"Sheet with name '{day_name}' not found in the workbook.")
      sheet = workbook[day_name]
      
      keywords = []
      for row in sheet.iter_rows(min_row=3, min_col=3, max_col=3, values_only=True):
            if row[0]:
                  keywords.append(row[0])
      return keywords, workbook, sheet




def search_google_and_get_suggestions(driver, keyword):
      """Searches for the given keyword on Google and returns the suggestions.

      Args:
          driver (Browser_instance): Opened browser instance.
          keyword (List_of_String): Keyword to search on Google.

      Returns:
          List: Returns the list of suggestions.
      """
           
      driver.get("https://www.google.com")
      time.sleep(2)
      
      search_box = driver.find_element(By.NAME, "q")
      search_box.send_keys(keyword)
      time.sleep(2)
      
      suggestions = driver.find_elements(By.CSS_SELECTOR, "ul.G43f7e li div.wM6W7d span")
      options = [suggestion.text for suggestion in suggestions if suggestion.text]
      search_box.send_keys(Keys.ESCAPE)
      return options




def find_longest_and_shortest(options):
      """Finds the longest and shortest strings from the given list of strings.

      Args:
          options (List): List of strings to find the longest and shortest strings.

      Returns:
          String: Returns the longest and shortest strings.
      """      
      if not options:
            return "", ""
      
      longest = max(options, key=len)
      shortest = min(options, key=len)
      
      print(f"Longest length: {len(longest)}")
      print(f"Shortest length: {len(shortest)}")
      
      return longest, shortest




def write_to_excel(sheet, row, longest, shortest):
      """Writes the longest and shortest strings to the excel sheet.

      Args:
          sheet (Instance_of_file): Excel sheet instance.
          row (Integer): Row number to write the data.
          longest (String): Longest string.
          shortest (String): Shortest string.
      """      
      sheet.cell(row=row, column=4, value=longest)
      sheet.cell(row=row, column=5, value=shortest)
      


      
def main():
      """Main function to run the program.
      """      
      
      # Initial setup for web automation
      file_path = "keywords.xlsx"
      
      current_day = get_day_of_week()
      print(f"Today is {current_day}")
      
      driver = webdriver.Chrome()
      
      try:
            # Read the keywords from the excel file
            keywords, workbook, sheet = read_keywords_from_excel(file_path, current_day)
            print(f"Found {len(keywords)} keywords. Processing...")
            
            
            # Process each keyword
            for idx, keyword in enumerate(keywords, start=3):
                  print(f"Searching for: {keyword}")
                  suggestions = search_google_and_get_suggestions(driver, keyword)
                  longest, shortest = find_longest_and_shortest(suggestions)
                  print(f"Longest suggestion: {longest}")
                  print(f"Shortest suggestion: {shortest}")
                  
                  # Write to excel
                  write_to_excel(sheet, idx, longest, shortest)

            # Save the workbook
            workbook.save(file_path)
            print("Saved the workbook.")
      
      except Exception as e:
            print(f"An error occurred: {e}")
      finally:
            # Close the browser
            driver.quit()
            print("Quitting the browser.")
            


            
if __name__ == "__main__":
      main()

